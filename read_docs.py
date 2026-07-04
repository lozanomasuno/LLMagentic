import docx
import openpyxl
import os

folder = 'C:/Users/lozan/OneDrive/Documentos/LLMagentic/Prueba Técnica v2.0'

docx_files = [f for f in os.listdir(folder) if f.endswith('.docx')]
for fname in sorted(docx_files):
    path = os.path.join(folder, fname)
    try:
        doc = docx.Document(path)
        text = '\n'.join([p.text for p in doc.paragraphs if p.text.strip()])
        print(f'\n{"="*60}')
        print(f'FILE: {fname}')
        print('='*60)
        print(text[:5000])
    except Exception as e:
        print(f'ERROR {fname}: {e}')

# Read Excel
xlsx_path = os.path.join(folder, 'BD_afiliados.xlsx')
try:
    wb = openpyxl.load_workbook(xlsx_path)
    print(f'\n{"="*60}')
    print(f'EXCEL: BD_afiliados.xlsx')
    print('='*60)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n--- Sheet: {sheet_name} ---')
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[:30]:
            print(row)
except Exception as e:
    print(f'ERROR Excel: {e}')
